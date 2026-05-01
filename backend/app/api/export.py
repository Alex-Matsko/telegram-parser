"""
Price list export endpoint.

Formats:
  - Excel (.xlsx)  — multi-sheet: one sheet per category + Summary sheet
  - CSV   (.csv)   — flat, all rows, UTF-8 BOM (Excel-friendly)
  - JSON  (.json)  — structured, grouped by category→brand→model

Endpoint:
  GET /api/price-list/export?format=xlsx&brand=Apple&condition=new&...
"""
import csv
import io
import json
from datetime import datetime, timezone
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_session
from app.services.price_service import get_price_list

router = APIRouter(prefix="/price-list", tags=["Export"])

# Memory sort order for consistent column arrangement
_MEMORY_ORDER = {
    "32GB": 0, "64GB": 1, "128GB": 2, "256GB": 3, "512GB": 4,
    "1TB": 5, "2TB": 6,
}


def _memory_sort_key(mem: Optional[str]) -> int:
    if mem is None:
        return 99
    return _MEMORY_ORDER.get(mem.upper(), 50)


def _decimal_to_float(val) -> Optional[float]:
    if val is None:
        return None
    return float(val)


def _ts() -> str:
    """Timestamp string for filenames."""
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")


async def _fetch_all_items(session: AsyncSession, filters: dict) -> list:
    """Fetch ALL offers ignoring pagination (per_page=5000)."""
    result = await get_price_list(
        session=session,
        brand=filters.get("brand"),
        model=filters.get("model"),
        memory=filters.get("memory"),
        color=filters.get("color"),
        condition=filters.get("condition"),
        supplier_id=filters.get("supplier_id"),
        currency=filters.get("currency"),
        price_min=filters.get("price_min"),
        price_max=filters.get("price_max"),
        updated_after=filters.get("updated_after"),
        sort_by=filters.get("sort_by", "brand"),
        order=filters.get("order", "asc"),
        page=1,
        per_page=5000,
    )
    return result.items


def _items_to_rows(items) -> list[dict]:
    """Convert PriceListItem objects to flat dicts for export."""
    rows = []
    for item in items:
        rows.append({
            "category": item.category or "",
            "brand": item.brand or "",
            "model": item.model or "",
            "memory": item.memory or "",
            "color": item.color or "",
            "condition": item.condition or "new",
            "sim_type": item.sim_type or "",
            "best_price": _decimal_to_float(item.best_price),
            "best_supplier": item.best_supplier or "",
            "second_price": _decimal_to_float(item.second_price),
            "second_supplier": item.second_supplier or "",
            "third_price": _decimal_to_float(item.third_price),
            "third_supplier": item.third_supplier or "",
            "offer_count": item.offer_count,
            "price_change_3d": _decimal_to_float(item.price_change_3d),
            "price_change_3d_pct": item.price_change_3d_pct,
            "last_updated": item.last_updated.strftime("%Y-%m-%d %H:%M") if item.last_updated else "",
        })
    return rows


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

CSV_HEADERS = [
    "category", "brand", "model", "memory", "color", "condition", "sim_type",
    "best_price", "best_supplier",
    "second_price", "second_supplier",
    "third_price", "third_supplier",
    "offer_count", "price_change_3d", "price_change_3d_pct", "last_updated",
]

CSV_HEADERS_RU = [
    "Категория", "Бренд", "Модель", "Память", "Цвет", "Состояние", "SIM",
    "Лучшая цена", "Поставщик 1",
    "Цена 2", "Поставщик 2",
    "Цена 3", "Поставщик 3",
    "Кол-во предл.", "Изм. цены 3д (руб)", "Изм. цены 3д (%)", "Обновлено",
]


def _build_csv(rows: list[dict]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(
        buf,
        fieldnames=CSV_HEADERS,
        extrasaction="ignore",
        lineterminator="\r\n",
    )
    # Russian header row
    writer.writerow(dict(zip(CSV_HEADERS, CSV_HEADERS_RU)))
    for row in rows:
        writer.writerow(row)
    # UTF-8 BOM so Excel opens without encoding issues
    return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# JSON export
# ---------------------------------------------------------------------------

def _build_json(rows: list[dict]) -> bytes:
    """Group by category → brand → model → list of variants."""
    grouped: dict = {}
    for row in rows:
        cat = row["category"] or "other"
        brand = row["brand"] or "Unknown"
        model = row["model"] or "Unknown"
        grouped.setdefault(cat, {}).setdefault(brand, {}).setdefault(model, []).append({
            "memory": row["memory"],
            "color": row["color"],
            "condition": row["condition"],
            "sim_type": row["sim_type"],
            "best_price": row["best_price"],
            "best_supplier": row["best_supplier"],
            "second_price": row["second_price"],
            "second_supplier": row["second_supplier"],
            "third_price": row["third_price"],
            "third_supplier": row["third_supplier"],
            "offer_count": row["offer_count"],
            "last_updated": row["last_updated"],
        })
    # Sort variants by memory inside each model
    for cat in grouped:
        for brand in grouped[cat]:
            for model in grouped[cat][brand]:
                grouped[cat][brand][model].sort(
                    key=lambda x: _memory_sort_key(x.get("memory"))
                )
    output = {
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "total_items": len(rows),
        "data": grouped,
    }
    return json.dumps(output, ensure_ascii=False, indent=2).encode("utf-8")


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _build_xlsx(rows: list[dict]) -> bytes:
    """
    Multi-sheet Excel:
      - Sheet "Все товары" — full flat table
      - One sheet per category (smartphone → iPhone/Samsung/..., etc.)
    Requires openpyxl (add to requirements.txt).
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Alignment, Font, PatternFill, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Add 'openpyxl>=3.1.0' to backend/requirements.txt and rebuild."
        )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # ---- colour palette ----
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")   # dark blue header
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    CAT_FILL = PatternFill("solid", fgColor="2E75B6")   # category sub-header
    CAT_FONT = Font(bold=True, color="FFFFFF", size=10)
    BRAND_FILL = PatternFill("solid", fgColor="D6E4F0")  # brand grouping row
    BRAND_FONT = Font(bold=True, color="1F4E79", size=10)
    PRICE_GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")   # green tint = cheapest
    EVEN_FILL = PatternFill("solid", fgColor="F5F9FF")   # alternating row

    PRICE_FMT = '#,##0\ ₽'
    PCT_FMT = '+0.0%;-0.0%;0.0%'

    def _col_widths():
        return [
            ("Категория", 16), ("Бренд", 14), ("Модель", 28), ("Память", 9),
            ("Цвет", 18), ("Состояние", 12), ("SIM", 10),
            ("Лучшая цена", 14), ("Поставщик 1", 20),
            ("Цена 2", 12), ("Поставщик 2", 20),
            ("Цена 3", 12), ("Поставщик 3", 20),
            ("Предл.", 8), ("Изм. 3д ₽", 12), ("Изм. 3д %", 10), ("Обновлено", 16),
        ]

    def _write_sheet(ws, sheet_rows: list[dict]):
        headers = [h for h, _ in _col_widths()]
        col_names = CSV_HEADERS

        # Header row
        for col_idx, (hdr, width) in enumerate(_col_widths(), start=1):
            cell = ws.cell(row=1, column=col_idx, value=hdr)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

        # Group rows by brand for alternating colours
        current_brand = None
        brand_toggle = False
        row_idx = 2
        sorted_rows = sorted(
            sheet_rows,
            key=lambda r: (
                r.get("category", ""),
                r.get("brand", ""),
                r.get("model", ""),
                _memory_sort_key(r.get("memory")),
            )
        )
        for data_row in sorted_rows:
            if data_row.get("brand") != current_brand:
                current_brand = data_row.get("brand")
                brand_toggle = not brand_toggle

            fill = EVEN_FILL if brand_toggle else PatternFill("solid", fgColor="FFFFFF")

            for col_idx, key in enumerate(col_names, start=1):
                val = data_row.get(key)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = fill
                cell.alignment = Alignment(vertical="center")

                # Format price columns
                if key in ("best_price", "second_price", "third_price", "price_change_3d"):
                    if val is not None:
                        cell.number_format = PRICE_FMT
                    # Highlight best price cell with green
                    if key == "best_price" and val is not None:
                        cell.fill = PRICE_GOOD_FILL
                elif key == "price_change_3d_pct":
                    if val is not None:
                        cell.number_format = PCT_FMT
                        cell.value = val / 100 if val else val  # openpyxl % needs fraction

            row_idx += 1

        # Auto-filter on header row
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_col_widths()))}1"

    # ---- Sheet 1: All items ----
    ws_all = wb.create_sheet("Все товары")
    _write_sheet(ws_all, rows)

    # ---- Per-category sheets ----
    categories = sorted({r.get("category", "other") for r in rows})
    CAT_NAMES_RU = {
        "smartphone": "Смартфоны",
        "laptop": "Ноутбуки",
        "tablet": "Планшеты",
        "headphones": "Наушники",
        "watch": "Умные часы",
        "camera": "Камеры",
        "console": "Консоли",
        "desktop": "Десктопы",
        "accessory": "Аксессуары",
        "vr": "VR",
        "tv": "ТВ",
        "appliance": "Техника",
    }
    for cat in categories:
        cat_rows = [r for r in rows if r.get("category") == cat]
        if not cat_rows:
            continue
        sheet_name = CAT_NAMES_RU.get(cat, cat.capitalize())[:31]  # Excel 31-char limit
        ws_cat = wb.create_sheet(sheet_name)
        _write_sheet(ws_cat, cat_rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Route
# ---------------------------------------------------------------------------

@router.get("/export", summary="Export price list (xlsx / csv / json)")
async def export_price_list(
    format: str = Query("xlsx", pattern="^(xlsx|csv|json)$", description="Export format"),
    brand: Optional[str] = Query(None),
    model: Optional[str] = Query(None),
    memory: Optional[str] = Query(None),
    color: Optional[str] = Query(None),
    condition: Optional[str] = Query(None),
    supplier_id: Optional[int] = Query(None),
    currency: Optional[str] = Query(None),
    price_min: Optional[float] = Query(None, ge=0),
    price_max: Optional[float] = Query(None, ge=0),
    updated_after: Optional[datetime] = Query(None),
    sort_by: str = Query("brand", pattern="^(best_price|model|brand|offer_count|last_updated)$"),
    order: str = Query("asc", pattern="^(asc|desc)$"),
    session: AsyncSession = Depends(get_session),
):
    """
    Export the current price list in the requested format.

    All filter parameters are identical to GET /price-list.
    Returns a downloadable file.
    """
    filters = dict(
        brand=brand, model=model, memory=memory, color=color,
        condition=condition, supplier_id=supplier_id, currency=currency,
        price_min=price_min, price_max=price_max, updated_after=updated_after,
        sort_by=sort_by, order=order,
    )
    items = await _fetch_all_items(session, filters)
    rows = _items_to_rows(items)
    ts = _ts()

    if format == "csv":
        data = _build_csv(rows)
        filename = f"pricelist_{ts}.csv"
        media_type = "text/csv; charset=utf-8-sig"
    elif format == "json":
        data = _build_json(rows)
        filename = f"pricelist_{ts}.json"
        media_type = "application/json"
    else:  # xlsx
        data = _build_xlsx(rows)
        filename = f"pricelist_{ts}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        io.BytesIO(data),
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(data)),
        },
    )
